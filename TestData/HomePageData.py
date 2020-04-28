import openpyxl


class HomePageData:
    test_HomePage_data = [{"firstName": "ritesh", "lastName": "mishra", "gender": "Male"},
                          {"firstName": "ranjan", "lastName": "mishra", "gender": "Male"}]

    @staticmethod
    def getTestData():

        Dict = {}
        workbook = openpyxl.load_workbook("D:\\PYTHON\\DataDriven.xlsx")

        sheet = workbook.active

        rows = sheet.max_row

        columns = sheet.max_column

        for i in range(2, rows + 1):

            for j in range(2, columns + 1):
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

            return [Dict]
